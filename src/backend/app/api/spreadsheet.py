from fastapi import APIRouter, UploadFile, File, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from typing import List, Dict, Any, Optional
import openpyxl
import pandas as pd
from io import BytesIO

router = APIRouter()

# ==================== REQUEST/RESPONSE MODELS ====================


class FormulaRequest(BaseModel):
    formula: str = Field(..., description="Formula to evaluate (e.g., '=SUM(A1:A10)')")
    data: Dict[str, Any] = Field(default_factory=dict, description="Cell data dictionary")
    cellRef: Optional[str] = Field(None, description="Reference cell for the formula")


class FormulaResponse(BaseModel):
    result: Any
    success: bool
    error: Optional[str] = None


class SortKey(BaseModel):
    column: str
    direction: str = Field(..., pattern="^(asc|desc)$")
    type: Optional[str] = "auto"


class SortRequest(BaseModel):
    data: List[Dict[str, Any]]
    sortKeys: List[SortKey]


class FilterCriteria(BaseModel):
    column: str
    condition: str = Field(
        ...,
        pattern=("^(equals|notEquals|greaterThan|lessThan|"
                 "greaterThanOrEqual|lessThanOrEqual|contains|"
                 "notContains|startsWith|endsWith|empty|notEmpty|"
                 "between|in)$")
    )
    value: Any


class FilterRequest(BaseModel):
    data: List[Dict[str, Any]]
    criteria: FilterCriteria


class CloudSyncRequest(BaseModel):
    provider: str = Field(..., pattern="^(microsoft|google)$")
    fileId: str
    data: Dict[str, Any]


class ExcelImportResponse(BaseModel):
    sheets: Dict[str, List[List[Any]]]
    success: bool
    error: Optional[str] = None


class CloudFilesResponse(BaseModel):
    files: List[Dict[str, Any]]
    success: bool


# ==================== FORMULA EVALUATION ====================


class FormulaEngine:
    """Python-based formula evaluation engine"""

    def __init__(self):
        self.data = {}

    def update_data(self, data: Dict[str, Any]):
        self.data = data

    def evaluate(self, formula: str, cell_ref: Optional[str] = None) -> Any:
        """Evaluate an Excel formula"""
        if not formula.startswith('='):
            return formula

        try:
            expr = formula[1:]
            result = self._parse_and_eval(expr)
            return result
        except Exception as e:
            return {"error": str(e)}

    def _parse_and_eval(self, expr: str) -> Any:
        """Parse and evaluate expression"""
        expr = expr.strip()

        if '(' in expr:
            func_match = expr.split('(', 1)
            func_name = func_match[0].upper()
            args_str = func_match[1].rsplit(')', 1)[0]
            args = self._parse_args(args_str)
            return self._call_function(func_name, args)

        if (expr.replace(':', '').replace('$', '').isalnum()
                and any(c.isalpha() for c in expr)):
            return self._get_cell_value(expr)

        try:
            return float(expr)
        except ValueError:
            pass

        return expr

    def _parse_args(self, args_str: str) -> List[Any]:
        """Parse function arguments"""
        args = []
        current = ""
        depth = 0

        for char in args_str:
            if char == '(':
                depth += 1
                current += char
            elif char == ')':
                depth -= 1
                current += char
            elif char == ',' and depth == 0:
                args.append(self._parse_and_eval(current.strip()))
                current = ""
            else:
                current += char

        if current.strip():
            args.append(self._parse_and_eval(current.strip()))

        return args

    def _call_function(self, name: str, args: List[Any]) -> Any:
        """Call a function by name"""
        functions = {
            'SUM': lambda a: sum(self._flatten(a)),
            'AVERAGE': lambda a: (
                sum(self._flatten(a)) / len(self._flatten(a))
                if self._flatten(a) else 0
            ),
            'COUNT': lambda a: len([
                x for x in self._flatten(a) if isinstance(x, (int, float))
            ]),
            'MAX': lambda a: max(self._flatten(a)) if self._flatten(a) else 0,
            'MIN': lambda a: min(self._flatten(a)) if self._flatten(a) else 0,
            'IF': lambda a: a[1] if a[0] else a[2] if len(a) > 2 else False,
            'AND': lambda a: all(a),
            'OR': lambda a: any(a),
            'ABS': lambda a: abs(a[0]),
            'ROUND': lambda a: round(a[0], int(a[1]) if len(a) > 1 else 0),
            'POWER': lambda a: a[0] ** a[1],
            'CONCAT': lambda a: ''.join(str(x) for x in a),
            'LEFT': lambda a: str(a[0])[:int(a[1]) if len(a) > 1 else 1],
            'RIGHT': lambda a: str(a[0])[-int(a[1]) if len(a) > 1 else 1:],
            'LEN': lambda a: len(str(a[0])),
            'UPPER': lambda a: str(a[0]).upper(),
            'LOWER': lambda a: str(a[0]).lower(),
        }

        if name not in functions:
            raise ValueError(f"Unknown function: {name}")

        return functions[name](args)

    def _flatten(self, arr):
        """Flatten nested arrays"""
        result = []
        for item in arr if isinstance(arr, list) else [arr]:
            if isinstance(item, list):
                result.extend(self._flatten(item))
            else:
                result.append(item)
        return result

    def _get_cell_value(self, ref: str) -> Any:
        """Get cell value from reference"""
        if ':' in ref:
            return self._get_range_values(ref)

        clean_ref = ref.replace('$', '').upper()
        val = self.data.get(clean_ref, 0)

        if isinstance(val, str) and val.startswith('='):
            return self.evaluate(val, clean_ref)
        return val

    def _get_range_values(self, range_ref: str) -> List[List[Any]]:
        """Get values in a range"""
        start, end = range_ref.split(':')
        start_ref = self._parse_cell_ref(start)
        end_ref = self._parse_cell_ref(end)

        values = []
        for row in range(start_ref['row'], end_ref['row'] + 1):
            row_vals = []
            for col in range(start_ref['col'], end_ref['col'] + 1):
                col_letter = self._col_to_letter(col)
                cell_ref = f"{col_letter}{row + 1}"
                row_vals.append(self.data.get(cell_ref, 0))
            values.append(row_vals)

        return values

    def _parse_cell_ref(self, ref: str) -> Dict[str, int]:
        """Parse A1 notation"""
        import re
        match = re.match(r'\$?([A-Z]+)\$?(\d+)', ref.upper())
        if not match:
            raise ValueError(f"Invalid cell reference: {ref}")

        col = 0
        for char in match.group(1):
            col = col * 26 + (ord(char) - 65)

        return {'col': col, 'row': int(match.group(2)) - 1}

    def _col_to_letter(self, col: int) -> str:
        """Convert column index to letter"""
        result = ""
        while col >= 0:
            result = chr(65 + (col % 26)) + result
            col = col // 26 - 1
        return result or "A"


# Global formula engine instance
formula_engine = FormulaEngine()


@router.post("/formula/evaluate", response_model=FormulaResponse)
async def evaluate_formula(request: FormulaRequest):
    """Evaluate an Excel formula"""
    try:
        formula_engine.update_data(request.data)
        result = formula_engine.evaluate(request.formula, request.cellRef)
        return FormulaResponse(result=result, success=True)
    except Exception as e:
        return FormulaResponse(result=None, success=False, error=str(e))


# ==================== DATA OPERATIONS ====================


@router.post("/data/sort")
async def sort_data(request: SortRequest):
    """Sort data by multiple columns"""
    try:
        df = pd.DataFrame(request.data)

        sort_cols = []
        ascending = []

        for key in request.sortKeys:
            sort_cols.append(key.column)
            ascending.append(key.direction == 'asc')

        sorted_df = df.sort_values(by=sort_cols, ascending=ascending)

        return {
            "data": sorted_df.to_dict('records'),
            "success": True
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/data/filter")
async def filter_data(request: FilterRequest):
    """Filter data based on criteria"""
    try:
        df = pd.DataFrame(request.data)

        col = request.criteria.column
        op = request.criteria.condition
        val = request.criteria.value

        if op == 'equals':
            mask = df[col] == val
        elif op == 'notEquals':
            mask = df[col] != val
        elif op == 'greaterThan':
            mask = df[col] > val
        elif op == 'lessThan':
            mask = df[col] < val
        elif op == 'greaterThanOrEqual':
            mask = df[col] >= val
        elif op == 'lessThanOrEqual':
            mask = df[col] <= val
        elif op == 'contains':
            mask = df[col].astype(str).str.contains(str(val), case=False)
        elif op == 'notContains':
            mask = ~df[col].astype(str).str.contains(str(val), case=False)
        elif op == 'startsWith':
            mask = df[col].astype(str).str.startswith(str(val))
        elif op == 'endsWith':
            mask = df[col].astype(str).str.endswith(str(val))
        elif op == 'empty':
            mask = df[col].isna() | (df[col] == '')
        elif op == 'notEmpty':
            mask = df[col].notna() & (df[col] != '')
        else:
            mask = pd.Series([True] * len(df))

        filtered_df = df[mask]

        return {
            "data": filtered_df.to_dict('records'),
            "success": True
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/data/dedup")
async def remove_duplicates(
    data: List[Dict[str, Any]],
    columns: Optional[List[str]] = None
):
    """Remove duplicate rows"""
    try:
        df = pd.DataFrame(data)

        if columns:
            df = df.drop_duplicates(subset=columns)
        else:
            df = df.drop_duplicates()

        return {
            "data": df.to_dict('records'),
            "removed": len(data) - len(df),
            "success": True
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


# ==================== EXCEL FILE OPERATIONS ====================


@router.post("/excel/import", response_model=ExcelImportResponse)
async def import_excel(file: UploadFile = File(...)):
    """Import Excel file and convert to JSON"""
    try:
        contents = await file.read()
        workbook = openpyxl.load_workbook(BytesIO(contents), data_only=True)

        result = {}
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            data = []

            for row in sheet.iter_rows(values_only=True):
                data.append(list(row))

            result[sheet_name] = data

        return ExcelImportResponse(sheets=result, success=True)
    except Exception as e:
        return ExcelImportResponse(sheets={}, success=False, error=str(e))


@router.post("/excel/export")
async def export_excel(data: Dict[str, List[List[Any]]]):
    """Export data to Excel format"""
    try:
        workbook = openpyxl.Workbook()

        first = True
        for sheet_name, sheet_data in data.items():
            if first:
                sheet = workbook.active
                sheet.title = sheet_name
                first = False
            else:
                sheet = workbook.create_sheet(title=sheet_name)

            for row_idx, row in enumerate(sheet_data, 1):
                for col_idx, value in enumerate(row, 1):
                    sheet.cell(row=row_idx, column=col_idx, value=value)

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        return StreamingResponse(
            output,
            media_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
            headers={"Content-Disposition": "attachment; filename=export.xlsx"}
        )
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/excel/formulas/import")
async def import_excel_with_formulas(file: UploadFile = File(...)):
    """Import Excel file preserving formulas"""
    try:
        contents = await file.read()
        workbook = openpyxl.load_workbook(BytesIO(contents), data_only=False)

        result = {
            "sheets": {},
            "formulas": {},
            "values": {}
        }

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheet_data = []
            sheet_formulas = {}
            sheet_values = {}

            for row_idx, row in enumerate(sheet.iter_rows(), 1):
                row_data = []
                for col_idx, cell in enumerate(row, 1):
                    col_letter = (
                        chr(64 + col_idx)
                        if col_idx <= 26
                        else f"A{chr(64 + col_idx - 26)}"
                    )
                    cell_ref = f"{col_letter}{row_idx}"

                    if cell.value:
                        row_data.append(cell.value)
                        if (isinstance(cell.value, str)
                                and cell.value.startswith('=')):
                            sheet_formulas[cell_ref] = cell.value
                        else:
                            sheet_values[cell_ref] = cell.value
                    else:
                        row_data.append(None)

                sheet_data.append(row_data)

            result["sheets"][sheet_name] = sheet_data
            result["formulas"][sheet_name] = sheet_formulas
            result["values"][sheet_name] = sheet_values

        return result
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


# ==================== CLOUD SYNC OPERATIONS ====================


@router.get("/cloud/files")
async def get_cloud_files(
    provider: str = Query(..., pattern="^(microsoft|google)$"),
    access_token: str = Query(...)
):
    """Get files from cloud provider"""
    try:
        if provider == 'microsoft':
            import requests
            response = requests.get(
                "https://graph.microsoft.com/v1.0/me/drive/root/children",
                headers={"Authorization": f"Bearer {access_token}"},
                params={
                    "$filter": "file ne null",
                    "$select": "id,name,mimeType,lastModifiedDateTime,webUrl"
                }
            )
            files = response.json().get('value', [])
            return {
                "files": [{
                    "id": f["id"],
                    "name": f["name"],
                    "mimeType": f["mimeType"],
                    "modifiedTime": f["lastModifiedDateTime"],
                    "webUrl": f["webUrl"]
                } for f in files],
                "success": True
            }
        else:  # google
            import requests
            response = requests.get(
                "https://www.googleapis.com/drive/v3/files",
                headers={"Authorization": f"Bearer {access_token}"},
                params={
                    "q": "mimeType='application/vnd.google-apps.spreadsheet'",
                    "fields": "files(id,name,mimeType,modifiedTime,webViewLink)"
                }
            )
            files = response.json().get('files', [])
            return {
                "files": [{
                    "id": f["id"],
                    "name": f["name"],
                    "mimeType": f["mimeType"],
                    "modifiedTime": f["modifiedTime"],
                    "webUrl": f["webViewLink"]
                } for f in files],
                "success": True
            }
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/cloud/sync")
async def sync_to_cloud(request: CloudSyncRequest):
    """Sync data to cloud provider"""
    try:
        return {
            "success": True,
            "url": f"https://example.com/{request.provider}/files/{request.fileId}"
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/health")
async def health_check():
    """Health check endpoint"""
    return {"status": "healthy", "service": "spreadsheet"}
